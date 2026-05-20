#!/usr/bin/env python3
"""Scrape X/Twitter users matching trading-related keywords via Apify, export to Excel.

Uses the apidojo/tweet-scraper actor to fetch recent tweets matching the given
search terms, then deduplicates the authors and writes a styled .xlsx file
sorted by follower count.

Requires APIFY_TOKEN in the environment.
"""

import argparse
import json
import os
import sys
import urllib.request
import urllib.error

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

APIFY_ACTOR = "apidojo~tweet-scraper"
APIFY_ENDPOINT = (
    f"https://api.apify.com/v2/acts/{APIFY_ACTOR}/run-sync-get-dataset-items"
)

DEFAULT_KEYWORDS = [
    "prop firm",
    "funded trader",
    "crypto trader",
    "FTMO",
    "MyForexFunds",
    "crypto signals",
    "forex signals",
    "funded account",
]

HEADERS = [
    "Handle",
    "Display Name",
    "Profile URL",
    "Followers",
    "Following",
    "Verified",
    "Bio",
    "Matched Keyword",
]
COLUMN_WIDTHS = [20, 28, 38, 12, 12, 10, 60, 22]


def search_tweets(token: str, keywords: list[str], max_items: int) -> list[dict]:
    """Call Apify's tweet-scraper synchronously and return the dataset items."""
    payload = {
        "searchTerms": keywords,
        "maxItems": max_items,
        "sort": "Latest",
        "tweetLanguage": "en",
    }
    url = f"{APIFY_ENDPOINT}?token={token}"
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=600) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        print(f"Apify HTTP {e.code}: {body}", file=sys.stderr)
        sys.exit(1)
    except urllib.error.URLError as e:
        print(f"Network error contacting Apify: {e.reason}", file=sys.stderr)
        sys.exit(1)


def extract_users(tweets: list[dict]) -> list[dict]:
    """Dedupe authors by handle. First occurrence wins for matched keyword."""
    users: dict[str, dict] = {}
    for tweet in tweets:
        author = tweet.get("author") or {}
        handle = author.get("userName")
        if not handle or handle in users:
            continue
        bio = (author.get("description") or "").replace("\n", " ").strip()
        users[handle] = {
            "handle": f"@{handle}",
            "name": author.get("name") or "",
            "url": f"https://x.com/{handle}",
            "followers": int(author.get("followers") or 0),
            "following": int(author.get("following") or 0),
            "verified": bool(
                author.get("isBlueVerified") or author.get("verified")
            ),
            "bio": bio,
            "keyword": tweet.get("searchTerm") or tweet.get("_searchTerm") or "",
        }
    return list(users.values())


def write_xlsx(users: list[dict], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Traders"

    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    users_sorted = sorted(users, key=lambda u: u["followers"], reverse=True)
    for u in users_sorted:
        ws.append(
            [
                u["handle"],
                u["name"],
                u["url"],
                u["followers"],
                u["following"],
                "Yes" if u["verified"] else "No",
                u["bio"],
                u["keyword"],
            ]
        )
        url_cell = ws.cell(row=ws.max_row, column=3)
        url_cell.hyperlink = u["url"]
        url_cell.font = Font(color="0563C1", underline="single")

    for i, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    ws.freeze_panes = "A2"

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--keywords",
        default=",".join(DEFAULT_KEYWORDS),
        help="Comma-separated search keywords",
    )
    parser.add_argument(
        "--max-tweets",
        type=int,
        default=200,
        help="Max tweets to fetch from Apify (cost dial)",
    )
    parser.add_argument(
        "--output", default="traders.xlsx", help="Output .xlsx path"
    )
    args = parser.parse_args()

    token = os.environ.get("APIFY_TOKEN")
    if not token:
        print("Error: APIFY_TOKEN environment variable not set.", file=sys.stderr)
        print(
            "Get a token at https://console.apify.com/account/integrations",
            file=sys.stderr,
        )
        sys.exit(1)

    keywords = [k.strip() for k in args.keywords.split(",") if k.strip()]
    if not keywords:
        print("Error: no keywords provided.", file=sys.stderr)
        sys.exit(1)

    print(f"Searching {len(keywords)} keyword(s): {keywords}")
    print(f"Fetching up to {args.max_tweets} tweets via Apify...")

    tweets = search_tweets(token, keywords, args.max_tweets)
    print(f"Received {len(tweets)} tweets")

    users = extract_users(tweets)
    print(f"Extracted {len(users)} unique users")

    if not users:
        print("No users found — try broader keywords or a higher --max-tweets.")
        sys.exit(0)

    write_xlsx(users, args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
